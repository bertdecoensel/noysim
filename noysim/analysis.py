# Noysim -- Noise simulation tools for Aimsun.
# Copyright (c) 2010-2011 by Bert De Coensel, Ghent University & Griffith University.
#
# Tools for preprocessing and analysis of Aimsun/Noysim simulation results

import numpy
import pylab

import openpyxl

import acoustics
from geo import Point, Direction
from emission import QLDCar as LightVehicle, QLDBDouble as HeavyVehicle


#---------------------------------------------------------------------------------------------------
# Loading of simulation results
#---------------------------------------------------------------------------------------------------

def loadLevels(filename):
  """ load the level timeseries at the different receivers from the supplied Excel file
      note: for the moment, only works with xlsx file format
  """
  # load workbook
  wb = openpyxl.reader.excel.load_workbook(filename)
  # load simulation parameters
  sheet = wb.get_sheet_by_name(name = 'simulation')
  for row in range(15):
    if str(sheet.cell(row = row, column = 0).value).startswith('Actually simulated'):
      duration = float(sheet.cell(row = row, column = 1).value)
    if str(sheet.cell(row = row, column = 0).value).startswith('Time step'):
      dt = float(sheet.cell(row = row, column = 1).value)
  nrows = int(round(duration/dt))
  # fetch the timeseries for each receiver
  sheet = wb.get_sheet_by_name(name = 'levels')
  result = []
  column = 1
  while str(sheet.cell(row = 0, column = column).value).startswith('Rcvr'):
    result.append(acoustics.TimeSeries([float(sheet.cell(row = i, column = column).value) for i in range(1, nrows+1)], dt=dt))
    column += 1
  return result


class Logger(object):
  """ class for loading and iterating logger files """
  def __init__(self, filename):
    object.__init__(self)
    self.filename = filename
    self.file = None # handle to the open file
    self.t = 0.0

  def createVehicle(self, tokens):
    """ create a Vehicle object based on the supplied logger line tokens """
    # fetch vehicle parameters
    vID = int(tokens[0])
    vLength = float(tokens[1])
    vPosition = Point(float(tokens[2]), float(tokens[3]))
    vDirection = Direction(bearing = float(tokens[4]))
    vSpeed = float(tokens[5])
    vAcceleration = float(tokens[6])
    # simple check between light and heavy vehicles
    vClass = {False: LightVehicle, True: HeavyVehicle}[(vLength>10.0)]
    return vClass(vid = vID, position = vPosition, direction = vDirection, speed = vSpeed, acceleration = vAcceleration)

  def __iter__(self):
    """ return the iterator """
    return self

  def next(self):
    """ return (t, vehicleList) for each timestep in the logger file """
    if self.t == None:
      # end of file was reached in previous step
      raise StopIteration
    if self.file == None:
      self.file = open(self.filename, 'r')
      self.t = float(self.file.readline().strip())
    vehicles = []
    currentTime = self.t
    # fetch all vehicles
    while True:
      tokens = self.file.readline().strip().split()
      if len(tokens) == 0:
        # end of file
        self.t = None
        self.file.close()
        break
      elif len(tokens) == 1:
        # end of timestep
        self.t = float(tokens[0])
        break
      else:
        # new vehicle
        vehicles.append(self.createVehicle(tokens))
    return (currentTime, vehicles)


#---------------------------------------------------------------------------------------------------
# Analyzing noise events
#---------------------------------------------------------------------------------------------------

def plotMadmax(ts, threshold = 60.0, drop = 5.0, droptime = 25.0, mindt = 3.0, bins = range(55, 85)):
  """ perform the MadMax noise event algorithm and plot statistics of results """
  # perform MadMax algorithm
  events = ts.madmax(threshold = threshold, drop = drop, droptime = droptime, mindt = mindt)
  eTimes, eLevels = zip(*events)
  print 'Number of events:', len(events)
  # plot timeseries
  pylab.figure()
  pylab.subplot(2,1,1)
  ts.plot(color='green')
  peaks = acoustics.TimeSeries([100*(t in eTimes) for t in ts.times()], dt = ts.dt())
  peaks.plot(color='red')
  # calculate and plot histogram
  pylab.subplot(2,1,2)
  pylab.hist(eLevels, bins = bins)


#---------------------------------------------------------------------------------------------------
# Test code
#---------------------------------------------------------------------------------------------------

if __name__ == '__main__':
  pass

  try:
    pylab.show()
  except:
    pass
