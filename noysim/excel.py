# Noysim -- Noise simulation tools for Aimsun.
# Copyright (c) 2010-2011 by Bert De Coensel, Ghent University & Griffith University.
#
# Wrapper for outputting results to Excel

import xlwt # for xls format
import openpyxl # for xlsx format


#---------------------------------------------------------------------------------------------------
# Basic Excel interface
#---------------------------------------------------------------------------------------------------

class ExcelOutput(object):
  """ Interface for writing output to Excel files """
  def __init__(self):
    object.__init__(self)

  def canHaveColumnRanges(self):
    """ return True if column ranges can be used (e.g. 'A:B' instead of 'A1:B8') """
    raise NotImplementedError

  def createSheets(self, names):
    """ creates a list of sheets with the given names """
    raise NotImplementedError

  def setValue(self, sheet, row, column, value, format=None):
    """ sets the value of a field
        sheet: name of the sheet (string)
        row, column: 0-based integers
        value: the value (int, float or string)
        format: can be None, 'bold' or 'float'
    """
    raise NotImplementedError

  def setFormula(self, sheet, row, column, formula, format=None):
    """ same as setValue, but with formulas (with comma as argument separator and without leading '=' sign) """
    raise NotImplementedError

  def setColumnWidth(self, sheet, column, width):
    """ set the column width (as a number of characters) """
    raise NotImplementedError

  def save(self, filename):
    """ save the workbook to the given filename (without extension) """
    raise NotImplementedError


#---------------------------------------------------------------------------------------------------
# Output to classic xls format
#---------------------------------------------------------------------------------------------------

class ClassicExcelOutput(ExcelOutput):
  """ Interface for writing output to xls files """
  def __init__(self):
    ExcelOutput.__init__(self)
    self._wb = xlwt.Workbook()
    self._styles = {'bold': xlwt.easyxf('font: bold True;'), 'float': xlwt.easyxf(num_format_str = '0.00')}
    self._ws = {}

  def canHaveColumnRanges(self):
    return False

  def createSheets(self, names):
    for name in names:
      # check if sheet was already created
      assert not (name in self._ws)
      # create sheet and add handle
      self._ws[name] = self._wb.add_sheet(name)

  def setValue(self, sheet, row, column, value, format=None):
    if format == None:
      self._ws[sheet].write(row, column, value)
    else:
      self._ws[sheet].write(row, column, value, self._styles[format])

  def setFormula(self, sheet, row, column, formula, format=None):
    self.setValue(sheet=sheet, row=row, column=column, value=xlwt.Formula(formula), format=format)

  def setColumnWidth(self, sheet, column, width):
    self._ws[sheet].col(column).width = 256*width

  def save(self, filename):
    if not filename.endswith('.xls'):
      filename += '.xls'
    self._wb.save(filename)


#---------------------------------------------------------------------------------------------------
# Output to Excel 2007 OpenXML (xlsx) format
#---------------------------------------------------------------------------------------------------

class OpenXMLExcelOutput(ExcelOutput):
  """ Interface for writing output to xlsx files """
  def __init__(self):
    ExcelOutput.__init__(self)
    self._wb = openpyxl.workbook.Workbook()
    self._wb.remove_sheet(self._wb.get_active_sheet())
    self._ws = {}

  def canHaveColumnRanges(self):
    return True

  def createSheets(self, names):
    for name in names:
      # check if sheet was already created
      assert not (name in self._ws)
      # create sheet and add handle
      ws = self._wb.create_sheet()
      ws.title = name
      self._ws[name] = ws

  def setValue(self, sheet, row, column, value, format=None):
    c = self._ws[sheet].cell(row=row, column=column)
    c.value = value
    if format == 'bold':
      c.style.font.bold = True
    if format == 'float':
      c.style.number_format.format_code = '0.00'

  def setFormula(self, sheet, row, column, formula, format=None):
    self.setValue(sheet=sheet, row=row, column=column, value=('='+formula), format=format)

  def setColumnWidth(self, sheet, column, width):
    letter = openpyxl.cell.get_column_letter(column+1)
    cd = openpyxl.worksheet.ColumnDimension(letter)
    cd.width = width
    self._ws[sheet].column_dimensions[letter] = cd

  def save(self, filename):
    if not filename.endswith('.xlsx'):
      filename += '.xlsx'
    self._wb.save(filename)


#---------------------------------------------------------------------------------------------------
# Convenience function
#---------------------------------------------------------------------------------------------------

def ExcelWorkbook(extension = 'xlsx'):
  """ create an Excel workbook with the given file format """
  if extension == 'xls':
    return ClassicExcelOutput()
  elif extension == 'xlsx':
    return OpenXMLExcelOutput()
  else:
    raise Exception('unknown file format: "%s"' % str(extension))


#---------------------------------------------------------------------------------------------------
# Test code
#---------------------------------------------------------------------------------------------------

if __name__ == '__main__':

  for extension in ('xls', 'xlsx'):
    wb = ExcelWorkbook(extension)
    wb.createSheets(['alpha', 'beta'])
    message = 'This is a serious test'
    wb.setValue('alpha', 0, 0, message, 'bold')
    wb.setValue('alpha', 1, 0, 2)
    wb.setValue('alpha', 1, 1, 3.141592, 'float')
    wb.setValue('alpha', 2, 0, 4)
    wb.setValue('alpha', 2, 1, 5.987654, 'float')
    wb.setColumnWidth('alpha', 0, 2+len(message))
    wb.setValue('beta', 0, 0, 4)
    wb.setFormula('beta', 0, 1, '2.0*vlookup(A1,alpha!A2:B3,2,false)', 'float')
    wb.save('testfile')
